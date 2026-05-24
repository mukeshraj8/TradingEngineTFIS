from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter


TARGET_SHEETS = ("AB6 OS", "AB14", "AB15", "AB16")
WINDOW_ROWS = 80
WINDOW_COLUMNS = 40
UNRESOLVED_FIELDS = (
    "allowed_monthly_statuses",
    "option_type",
    "entry_time",
    "recalculation_time",
    "start_strike_formula",
    "end_strike_formula",
    "ideal_premium_formula",
    "minimum_premium_formula",
    "entry_formula",
    "target_formula",
    "stoploss_formula",
    "minimum_oi",
    "carry_forward_allowed",
)


@dataclass(frozen=True, slots=True)
class DiscoveryAnchor:
    value: str
    coordinate: str


@dataclass(slots=True)
class S23FormulaBlockDiscovery:
    workbook_path: Path
    strategy_code: str = "S23"
    unique_code: str = "NIFTY_OP_SELL_WK_DIFF_2D_3D"

    def discover(self) -> dict[str, Any]:
        workbook_path = Path(self.workbook_path)
        if not workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {workbook_path}")

        workbook_values = load_workbook(workbook_path, read_only=False, data_only=True)
        workbook_formulas = load_workbook(workbook_path, read_only=False, data_only=False)
        try:
            sheets_payload: dict[str, Any] = {}
            for sheet_name in TARGET_SHEETS:
                values_sheet = workbook_values[sheet_name]
                formulas_sheet = workbook_formulas[sheet_name]
                anchors = self._find_sheet_anchors(values_sheet)
                nearby_cells = self._collect_nearby_cells(
                    values_sheet=values_sheet,
                    formulas_sheet=formulas_sheet,
                    anchors=anchors,
                )
                sheets_payload[sheet_name] = {
                    "anchors": anchors,
                    "nearby_cells": nearby_cells,
                }

            field_candidates = self._build_field_candidates(sheets_payload)
            return {
                "workbook_path": str(workbook_path),
                "strategy_code": self.strategy_code,
                "unique_code": self.unique_code,
                "window": {
                    "rows": WINDOW_ROWS,
                    "columns": WINDOW_COLUMNS,
                },
                "sheets": sheets_payload,
                "field_candidates": field_candidates,
                "open_questions": self._open_questions(field_candidates),
            }
        finally:
            workbook_values.close()
            workbook_formulas.close()

    @staticmethod
    def write_json(data: dict[str, Any], out_path: str | Path) -> Path:
        path = Path(out_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(data, indent=2), encoding="utf-8")
        return path

    @staticmethod
    def write_markdown(discovery: dict[str, Any], out_path: str | Path) -> Path:
        path = Path(out_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        lines = [
            "# S23 Formula Block Discovery",
            "",
            f"Workbook: `{discovery['workbook_path']}`",
            "",
            "## Field Candidates",
        ]

        for field_name in UNRESOLVED_FIELDS:
            details = discovery["field_candidates"][field_name]
            lines.append(f"### `{field_name}`")
            lines.append(f"- Confidence: `{details['confidence']}`")
            if details["recommended_extraction_mapping"]:
                lines.append(
                    f"- Recommended extraction mapping: {details['recommended_extraction_mapping']}"
                )
            if details["open_questions"]:
                for question in details["open_questions"]:
                    lines.append(f"- Open question: {question}")
            if details["candidate_cells"]:
                lines.append("- Likely candidate cells:")
                for cell in details["candidate_cells"]:
                    value = cell["value"]
                    formula = cell["formula"]
                    lines.append(
                        f"  - `{cell['sheet']}!{cell['coordinate']}` value=`{value}` formula=`{formula}`"
                    )
            else:
                lines.append("- Likely candidate cells: none found")
            lines.append("")

        lines.extend(["## Sheet Anchors", ""])
        for sheet_name, sheet_details in discovery["sheets"].items():
            lines.append(f"### `{sheet_name}`")
            if sheet_details["anchors"]:
                for anchor in sheet_details["anchors"]:
                    lines.append(
                        f"- Anchor `{anchor['coordinate']}` = `{anchor['value']}`"
                    )
            else:
                lines.append("- No S23 anchors found on this sheet")
            lines.append(
                f"- Nearby non-empty cells captured: `{len(sheet_details['nearby_cells'])}`"
            )
            lines.append("")

        if discovery["open_questions"]:
            lines.append("## Open Questions")
            for question in discovery["open_questions"]:
                lines.append(f"- {question}")
            lines.append("")

        path.write_text("\n".join(lines), encoding="utf-8")
        return path

    def _find_sheet_anchors(self, worksheet: Any) -> list[dict[str, str]]:
        anchors: list[dict[str, str]] = []
        for row in worksheet.iter_rows():
            for cell in row:
                normalized = self._normalized_text(cell.value)
                if normalized in {self.strategy_code, self.unique_code}:
                    anchors.append(
                        {
                            "coordinate": cell.coordinate,
                            "value": normalized,
                        }
                    )
        return anchors

    def _collect_nearby_cells(
        self,
        *,
        values_sheet: Any,
        formulas_sheet: Any,
        anchors: list[dict[str, str]],
    ) -> list[dict[str, Any]]:
        coordinates: set[tuple[int, int]] = set()
        max_row = int(values_sheet.max_row or 0)
        max_column = int(values_sheet.max_column or 0)

        for anchor in anchors:
            row_index, column_index = coordinate_to_tuple(anchor["coordinate"])
            start_row = max(1, row_index - WINDOW_ROWS)
            end_row = min(max_row, row_index + WINDOW_ROWS)
            start_column = max(1, column_index - WINDOW_COLUMNS)
            end_column = min(max_column, column_index + WINDOW_COLUMNS)
            for current_row in range(start_row, end_row + 1):
                for current_column in range(start_column, end_column + 1):
                    coordinates.add((current_row, current_column))

        nearby_cells: list[dict[str, Any]] = []
        for row_index, column_index in sorted(coordinates):
            value = values_sheet.cell(row_index, column_index).value
            formula_value = formulas_sheet.cell(row_index, column_index).value
            if value in (None, "") and formula_value in (None, ""):
                continue
            coordinate = f"{get_column_letter(column_index)}{row_index}"
            serialized_value = self._serialize(value)
            serialized_formula = self._serialize(formula_value)
            normalized_text = self._normalized_text(value if value not in (None, "") else formula_value)
            nearby_cells.append(
                {
                    "sheet": values_sheet.title,
                    "coordinate": coordinate,
                    "value": serialized_value,
                    "formula": serialized_formula,
                    "normalized_text": normalized_text,
                    "row_index": row_index,
                    "column_index": column_index,
                    "tags": self._tags_for_text(normalized_text),
                }
            )
        return nearby_cells

    def _build_field_candidates(self, sheets_payload: dict[str, Any]) -> dict[str, Any]:
        index = self._index_cells(sheets_payload)

        def pick(*refs: tuple[str, str]) -> list[dict[str, Any]]:
            result: list[dict[str, Any]] = []
            for sheet_name, coordinate in refs:
                cell = index.get((sheet_name, coordinate))
                if cell is not None:
                    result.append(cell)
            return result

        def tag_pick(tag: str, *, preferred_sheet: str | None = None) -> list[dict[str, Any]]:
            matches: list[dict[str, Any]] = []
            for cell in index.values():
                if tag in cell["tags"] and (preferred_sheet is None or cell["sheet"] == preferred_sheet):
                    matches.append(cell)
            return sorted(matches, key=lambda item: (item["sheet"], item["row_index"], item["column_index"]))

        field_candidates = {
            "allowed_monthly_statuses": {
                "confidence": "high",
                "candidate_cells": pick(("AB6 OS", "D162"), ("AB6 OS", "D168")),
                "recommended_extraction_mapping": "Use `AB6 OS!D162` for Bull/Bull CF and `AB6 OS!D168` for Bear/Bear CF rule families.",
                "open_questions": [],
            },
            "option_type": {
                "confidence": "high",
                "candidate_cells": pick(("AB6 OS", "F162"), ("AB6 OS", "F165"), ("AB6 OS", "F168"), ("AB6 OS", "F171")),
                "recommended_extraction_mapping": "Use the option-type cell on the same AB6 OS rule row as the selected monthly-status branch.",
                "open_questions": [],
            },
            "entry_time": {
                "confidence": "medium",
                "candidate_cells": pick(("AB6 OS", "L175"), ("AB6 OS", "B176"), ("AB6 OS", "L178"), ("AB6 OS", "B180")),
                "recommended_extraction_mapping": "Prefer `AB6 OS!L175` for call-sell ORPT entry time; confirm whether the row-176 mirrored `09:24:59` cell is the canonical source.",
                "open_questions": [
                    "The workbook shows both ORPT rows and mirrored timing rows. The canonical entry-time source should be confirmed before YAML generation.",
                ],
            },
            "recalculation_time": {
                "confidence": "high",
                "candidate_cells": pick(("AB6 OS", "C176"), ("AB6 OS", "L176"), ("AB6 OS", "L177"), ("AB6 OS", "C180")),
                "recommended_extraction_mapping": "Use the AB6 OS recalculation rows (`176-180`) with `09:29:59` as the primary source for recalculation time.",
                "open_questions": [],
            },
            "start_strike_formula": {
                "confidence": "high",
                "candidate_cells": pick(("AB6 OS", "G162"), ("AB6 OS", "M176"), ("AB6 OS", "R183")),
                "recommended_extraction_mapping": "Use `AB6 OS!G162` for the base Bull/Call start-strike rule. Treat `AB6 OS!M176` and the `R183`-style gap rows as recalculation or position-open variants.",
                "open_questions": [],
            },
            "end_strike_formula": {
                "confidence": "high",
                "candidate_cells": pick(("AB6 OS", "G163"), ("AB6 OS", "O176"), ("AB6 OS", "S183")),
                "recommended_extraction_mapping": "Use `AB6 OS!G163` for the base Bull/Call end-strike rule; the later timing rows look like recalculation variants.",
                "open_questions": [],
            },
            "ideal_premium_formula": {
                "confidence": "high",
                "candidate_cells": pick(("AB6 OS", "H162"), ("AB6 OS", "T176"), ("AB6 OS", "U183")),
                "recommended_extraction_mapping": "Use `AB6 OS!H162` for the base ideal premium and keep the timing rows as recalculation variants.",
                "open_questions": [],
            },
            "minimum_premium_formula": {
                "confidence": "high",
                "candidate_cells": pick(("AB6 OS", "H163"), ("AB6 OS", "V176"), ("AB6 OS", "W183")),
                "recommended_extraction_mapping": "Use `AB6 OS!H163` for the base minimum premium and treat the later rows as timing-gap variants.",
                "open_questions": [],
            },
            "entry_formula": {
                "confidence": "high",
                "candidate_cells": pick(("AB6 OS", "M162"), ("AB6 OS", "O162"), ("AB14", "F48")),
                "recommended_extraction_mapping": "Use `AB6 OS!M162` as the Bull/Call entry formula; `AB6 OS!O162` is the first target expression tied to that entry.",
                "open_questions": [],
            },
            "target_formula": {
                "confidence": "high",
                "candidate_cells": pick(("AB6 OS", "O162"), ("AB6 OS", "O165"), ("AB6 OS", "O168"), ("AB6 OS", "O171")),
                "recommended_extraction_mapping": "Use the target column on AB6 OS (`O`) for each rule row; `AB6 OS!O162` is the Bull/Call target formula.",
                "open_questions": [],
            },
            "stoploss_formula": {
                "confidence": "high",
                "candidate_cells": pick(("AB6 OS", "M163"), ("AB6 OS", "M166"), ("AB6 OS", "M169"), ("AB6 OS", "M172")),
                "recommended_extraction_mapping": "Use the AB6 OS row immediately below each entry row for the paired SL/TRP formula; `AB6 OS!M163` is the Bull/Call stoploss formula.",
                "open_questions": [],
            },
            "minimum_oi": {
                "confidence": "high",
                "candidate_cells": pick(("AB6 OS", "I162"), ("AB6 OS", "I165"), ("AB6 OS", "I168"), ("AB6 OS", "I171")),
                "recommended_extraction_mapping": "Use the AB6 OS OI column (`I`) on the selected rule row. The current workbook text is `500 Lots`, so numeric normalization will still be needed later.",
                "open_questions": [
                    "The workbook stores OI as `500 Lots`, so a later normalization rule is needed to strip the unit text safely.",
                ],
            },
            "carry_forward_allowed": {
                "confidence": "medium",
                "candidate_cells": pick(("AB6 OS", "H160"), ("AB6 OS", "T160")),
                "recommended_extraction_mapping": "Tentatively map carry-forward permission from the `Yes` cells on the AB6 OS close-at-03:00 continuation rules.",
                "open_questions": [
                    "Carry-forward appears in operational note rows rather than the main rule table, so the canonical source should be confirmed.",
                ],
            },
        }

        # Add tag-driven fallbacks where coordinate lookups were sparse.
        if not field_candidates["allowed_monthly_statuses"]["candidate_cells"]:
            field_candidates["allowed_monthly_statuses"]["candidate_cells"] = tag_pick("BULL", preferred_sheet="AB6 OS")[:4]
        if not field_candidates["option_type"]["candidate_cells"]:
            field_candidates["option_type"]["candidate_cells"] = tag_pick("CALL", preferred_sheet="AB6 OS")[:4]

        return field_candidates

    @staticmethod
    def _index_cells(sheets_payload: dict[str, Any]) -> dict[tuple[str, str], dict[str, Any]]:
        index: dict[tuple[str, str], dict[str, Any]] = {}
        for sheet_details in sheets_payload.values():
            for cell in sheet_details["nearby_cells"]:
                index[(cell["sheet"], cell["coordinate"])] = cell
        return index

    @staticmethod
    def _tags_for_text(text: str) -> list[str]:
        normalized = text.upper()
        patterns = {
            "BULL": ("BULL",),
            "BULL_CF": ("BULL CF", "BULL / BULL CF"),
            "BEAR": ("BEAR",),
            "BEAR_CF": ("BEAR CF", "BEAR / BEAR CF"),
            "CALL": ("CALL",),
            "PUT": ("PUT",),
            "ENTRY": ("ENTRY",),
            "TGT": ("TGT",),
            "SL": ("SL",),
            "TRP": ("TRP",),
            "OI": ("OI",),
            "PREMIUM": ("PREMIUM",),
            "09:24:59": ("09:24:59",),
            "09:29:59": ("09:29:59",),
            "2DHH": ("2DHH",),
            "2DLL": ("2DLL",),
            "3DHH": ("3DHH",),
            "3DLL": ("3DLL",),
            "CE_ENTRY": ("CE : ENTRY",),
            "PE_ENTRY": ("PE : ENTRY",),
            "START_STRIKE": ("START STRIKE",),
            "END_STRIKE": ("END STRIKE",),
        }
        tags = [
            tag
            for tag, needles in patterns.items()
            if any(needle in normalized for needle in needles)
        ]
        return tags

    @staticmethod
    def _normalized_text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, time):
            return value.isoformat()
        return str(value).strip()

    @staticmethod
    def _serialize(value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, (datetime, date, time)):
            return value.isoformat()
        if isinstance(value, (int, float, bool, str)):
            return value
        return repr(value)

    @staticmethod
    def _open_questions(field_candidates: dict[str, Any]) -> list[str]:
        questions: list[str] = []
        for field_name, details in field_candidates.items():
            for question in details["open_questions"]:
                questions.append(f"{field_name}: {question}")
        return questions


def discover_s23_formula_blocks(workbook_path: str | Path) -> dict[str, Any]:
    discovery = S23FormulaBlockDiscovery(Path(workbook_path))
    return discovery.discover()
