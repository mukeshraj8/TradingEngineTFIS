from __future__ import annotations

from dataclasses import dataclass
import json
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook


_STRATEGY_CODE_PATTERN = re.compile(r"^S\d{1,3}$", re.IGNORECASE)


def _cell_reference(row: int, column: int) -> str:
    result = []
    current = column
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        result.append(chr(65 + remainder))
    return f"{''.join(reversed(result))}{row}"


@dataclass(slots=True)
class ExcelWorkbookProfiler:
    workbook_path: Path

    def profile(self) -> dict[str, Any]:
        workbook_path = Path(self.workbook_path)
        if not workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {workbook_path}")

        workbook = load_workbook(
            workbook_path,
            read_only=True,
            data_only=True,
        )
        try:
            sheets: list[dict[str, Any]] = []
            for worksheet in workbook.worksheets:
                max_row = int(worksheet.max_row or 0)
                max_column = int(worksheet.max_column or 0)
                non_empty_count = 0
                strategy_hits: list[dict[str, str]] = []
                unique_code_hits: list[dict[str, str]] = []

                for row in worksheet.iter_rows():
                    for cell in row:
                        value = cell.value
                        if value in (None, ""):
                            continue
                        non_empty_count += 1
                        text = str(value).strip()
                        if not text:
                            continue
                        if _STRATEGY_CODE_PATTERN.fullmatch(text):
                            strategy_hits.append(
                                {
                                    "cell": _cell_reference(cell.row, cell.column),
                                    "value": text,
                                }
                            )
                        if "_OP_" in text.upper():
                            unique_code_hits.append(
                                {
                                    "cell": _cell_reference(cell.row, cell.column),
                                    "value": text,
                                }
                            )

                sheets.append(
                    {
                        "name": worksheet.title,
                        "row_count": max_row,
                        "column_count": max_column,
                        "non_empty_cell_count": non_empty_count,
                        "likely_strategy_code_locations": strategy_hits,
                        "likely_unique_code_locations": unique_code_hits,
                    }
                )

            return {
                "workbook_path": str(workbook_path),
                "sheet_names": [sheet["name"] for sheet in sheets],
                "sheets": sheets,
            }
        finally:
            workbook.close()

    @staticmethod
    def write_json(profile: dict[str, Any], out_path: str | Path) -> Path:
        path = Path(out_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(profile, indent=2), encoding="utf-8")
        return path


def profile_workbook(workbook_path: str | Path, out_path: str | Path | None = None) -> dict[str, Any]:
    profiler = ExcelWorkbookProfiler(Path(workbook_path))
    profile = profiler.profile()
    if out_path is not None:
        profiler.write_json(profile, out_path)
    return profile
